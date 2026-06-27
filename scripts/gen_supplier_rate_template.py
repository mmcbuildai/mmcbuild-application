#!/usr/bin/env python3
"""
Generate the MMC Build supplier rate-card template (.xlsx).

A supplier-facing workbook for capturing REAL pricing to replace the AI
placeholder rates in MMC Quote. Tabbed by unit-of-measure type so a supplier
fills the rows that match how they actually quote (per m2, per lineal m, per
m3, per unit, or lump sum), and flags which construction SYSTEM each rate is
for (the whole point is getting MMC pricing, not just traditional).

Rows are pre-filled (category / element / unit, locked-look grey); the supplier
fills the yellow columns (rate, lead time, min qty, notes). Re-run any time:
    python scripts/gen_supplier_rate_template.py
Output: docs/MMC-Supplier-Rate-Template.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(os.path.dirname(__file__), "..", "docs", "MMC-Supplier-Rate-Template.xlsx")

NAVY = "1E293B"; TEAL = "0D9488"; GREY = "F1F5F9"; YELLOW = "FEF9C3"; AMBER = "B45309"
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=NAVY, size=15)
SUB_FONT = Font(color="475569", size=10)
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SYSTEMS = '"Traditional,Panelised,Volumetric,3D-printed,Any"'

# (cost category label, element/description, unit) grouped per UoM tab.
TABS = {
    "Area rates (per m2)": [
        ("Substructure", "Slab on ground — incl. excavation, formwork, mesh, pour", "m2 floor"),
        ("Frame", "Wall framing — timber stud (supply + erect)", "m2 wall"),
        ("Frame", "Floor framing — bearers/joists or cassette", "m2 floor"),
        ("Roof", "Roof structure + covering (trusses/cassette + sheet/tile)", "m2 roof"),
        ("External Walls & Cladding", "Brick veneer", "m2 wall"),
        ("External Walls & Cladding", "Lightweight cladding (FC / weatherboard)", "m2 wall"),
        ("External Walls & Cladding", "SIP / structural panel (supply + install)", "m2 wall"),
        ("Internal Walls & Partitions", "Stud + plasterboard partition", "m2 wall"),
        ("Wall Finishes", "Paint / render to walls", "m2 wall"),
        ("Floor Finishes", "Tiling", "m2 floor"),
        ("Floor Finishes", "Engineered timber / laminate", "m2 floor"),
        ("Floor Finishes", "Carpet", "m2 floor"),
        ("Ceiling Finishes", "Plasterboard ceiling + paint", "m2 ceiling"),
    ],
    "Lineal rates (per m)": [
        ("Substructure", "Strip / edge footing", "lineal m"),
        ("Fitments", "Skirting", "lineal m"),
        ("Fitments", "Cornice", "lineal m"),
        ("External Works", "Boundary fencing", "lineal m"),
        ("External Works", "Retaining wall", "lineal m"),
        ("Plumbing & Drainage", "In-ground drainage run", "lineal m"),
    ],
    "Volume rates (per m3)": [
        ("Substructure", "Concrete — footings", "m3"),
        ("Substructure", "Concrete — slab/raft", "m3"),
        ("External Works", "Imported fill / hardcore", "m3"),
    ],
    "Unit rates (each)": [
        ("Windows & External Doors", "Window unit (avg, supply + install)", "each"),
        ("Windows & External Doors", "External door (avg, supply + install)", "each"),
        ("Internal Doors", "Internal door (supply + hang)", "each"),
        ("Fitments", "Kitchen (supply + install, avg dwelling)", "each"),
        ("Fitments", "Bathroom vanity", "each"),
        ("Fitments", "Built-in wardrobe", "each"),
        ("Plumbing & Drainage", "Sanitary fixture — basin / toilet / shower (avg)", "each"),
        ("Plumbing & Drainage", "Hot water unit (supply + install)", "each"),
        ("Electrical", "GPO / power point", "No."),
        ("Electrical", "Light point", "No."),
        ("Electrical", "Switchboard", "each"),
        ("Mechanical (HVAC)", "Split-system AC (supply + install)", "each"),
        ("Fire Services", "Hardwired interconnected smoke alarm", "each"),
    ],
    "Lump sum & item": [
        ("Preliminaries", "Site establishment, fencing, amenities", "item"),
        ("Preliminaries", "Site supervision (per project)", "item"),
        ("Preliminaries", "Insurances + permits", "item"),
        ("Plumbing & Drainage", "Sewer / water service connection", "item"),
        ("Electrical", "Mains supply connection", "item"),
        ("Mechanical (HVAC)", "Ducted system (whole dwelling)", "item"),
        ("External Works", "Driveway + paths", "item"),
        ("External Works", "Landscaping allowance", "item"),
        ("Contingency", "Contingency (state basis: % or $)", "item / %"),
    ],
}

COLS = ["Cost category", "Element / description", "Unit", "Construction system",
        "Your rate (AUD, ex GST)", "Lead time (weeks)", "Min qty / order", "Notes"]
COL_W = [26, 52, 14, 18, 20, 16, 16, 40]

wb = Workbook()

# ── Read me tab ──
ws = wb.active
ws.title = "Read me"
ws.sheet_view.showGridLines = False
ws["A1"] = "MMC Build — Supplier Rate Card"; ws["A1"].font = TITLE_FONT
readme = [
    "",
    "Thank you for providing pricing. We're building a cost engine that compares traditional",
    "construction against modern methods (panelised, volumetric, 3D-printed), and we need real",
    "supplier rates to replace our placeholder figures.",
    "",
    "HOW TO FILL THIS IN",
    "  • Each tab is a unit-of-measure type. Fill only the rows that match how YOU quote.",
    "  • Grey columns (category / element / unit) are pre-filled — leave them as is.",
    "  • Fill the yellow columns: your rate, lead time, minimum order, and any notes.",
    "  • Construction system: pick which method the rate is for (drop-down). Use 'Any' if it",
    "    applies across methods. Add a row per system if your rate differs by method.",
    "  • All rates AUD, excluding GST. Note your rate basis (supply only / supply + install) in Notes.",
    "  • Add rows freely — the categories mirror our estimate structure but aren't exhaustive.",
    "",
    "RETURN TO",
    "  Karen Engel — karen.engel@mmcbuild.com.au",
    "",
    "Rates are indicative and used to calibrate comparative estimates; they are not a binding quote.",
]
for i, line in enumerate(readme, start=2):
    ws[f"A{i}"] = line
    ws[f"A{i}"].font = Font(bold=True, color=NAVY) if line.isupper() and line.strip() else SUB_FONT
ws.column_dimensions["A"].width = 100

# ── Rate tabs ──
for tab, rows in TABS.items():
    ws = wb.create_sheet(tab[:31])
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = f"MMC Build rate card — {tab}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:H2")
    ws["A2"] = "Fill the yellow columns. Pick the construction system per row. AUD ex GST."
    ws["A2"].font = SUB_FONT
    # header row (row 4)
    hr = 4
    for c, (name, w) in enumerate(zip(COLS, COL_W), start=1):
        cell = ws.cell(row=hr, column=c, value=name)
        cell.font = HEAD_FONT
        cell.fill = PatternFill("solid", fgColor=TEAL if c >= 5 else NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[chr(64 + c)].width = w
    ws.row_dimensions[hr].height = 30
    ws.freeze_panes = f"A{hr + 1}"
    # data rows
    dv = DataValidation(type="list", formula1=SYSTEMS, allow_blank=True)
    ws.add_data_validation(dv)
    for r, (cat, elem, unit) in enumerate(rows, start=hr + 1):
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=elem)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value="Any")
        dv.add(ws.cell(row=r, column=4))
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 2))
            if c <= 3:
                cell.fill = PatternFill("solid", fgColor=GREY)  # pre-filled, leave as-is
            elif c == 5:
                cell.fill = PatternFill("solid", fgColor=YELLOW)  # the rate — fill me
                cell.number_format = '#,##0.00'
        ws.row_dimensions[r].height = 22

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"wrote {os.path.relpath(OUT)} — {len(TABS)} rate tabs + Read me")
